from bs4 import BeautifulSoup
import requests 
import openpyxl


""""
excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'Naija in Foreign Media'
print(excel.sheetnames) 
sheet.append(['movie', 'movie_year', 'ratings'])
"""

movieList = []

try:
    source = requests.get('https://www.imdb.com/search/keyword/?keywords=nigeria&ref_=kw_vw_smp&sort=moviemeter,asc&mode=simple&page=1')
    #source.raise_for_status()

    soup = BeautifulSoup(source.text, 'html.parser')
    movies = soup.find_all('div', class_= "lister-item mode-simple")#.find_all('div', class_="lister-item mode-advanced")
    for movie in movies:
         title = movie.find('div', class_= "lister-item-content").a.text 
         movieYear = movie.find('span', class_= "lister-item-year text-muted unbold").text 
         try:
            ratings = movie.find('div', class_= "col-imdb-rating").strong.text
            print(ratings)
         except:
            ratings = movie.find('span', class_= "ghost")
            pass
    movieList.append(movie) 
    print(movieList)
        #genre =
except Exception as e:
    print(e)

